"""One-time build: parse every already-fetched SBI and Axis Mutual Fund
portfolio workbook (`data/amc_portfolios/{sbi,axis}/`, from `scripts/
fetch_amc_portfolios.py`) into ONE consolidated table of per-stock EQUITY
holdings - Tier 1 item 4's actual point. NOT a live/network call - pure
local file parsing, same category as `dtest/data/amc_holdings.py`'s own
scope (see that module's docstring for why equity-only, the shared
parser, and the two real file formats both named `.xls`).

    python scripts/build_amc_equity_holdings.py

AXIS: processes every manifest row with `scheme_name` set (the per-scheme
files - 3,611 after dedup) PLUS every real scheme sheet inside the ONE
consolidated workbook (2021-09-30) that predates per-scheme coverage.

SBI: each of the 68 monthly workbooks already covers every scheme as its
own sheet (SBI never split into per-scheme files at all) - scheme
identity for each sheet comes from that workbook's own "Index" sheet
(short-code -> code -> full name) where available, falling back to that
scheme's own identity rows (`_sbi_scheme_own_identity`) for the ~10 older
real files whose "Index" sheet is a bare list of sheet names with no
code/name columns at all.

REAL, MATERIAL GAP FOUND ONLY AFTER RUNNING THE FULL BUILD, NOT CAUGHT BY
THE EARLIER SAMPLE CHECKS: SBI's actual equity-holdings coverage in the
output is 2023-01 to 2026-07 (43 months), NOT the full 2013-2026 span its
own 68 workbooks nominally cover. The reason is a THIRD real template era
this session's inspection missed - SBI's 2013-2016 files (the same
"older naming era" `amc_portfolios.py` already documents for filenames)
also use a genuinely different, narrower HOLDINGS TABLE layout (a header
of `NAME OF THE INSTRUMENT, ISIN, QUANTITY, MARKET VALUE, RATING,
REMARKS` - 6 columns, no separate Industry/%-to-NAV split the same way
the modern template has) that `parse_scheme_sheet`'s 7-field equity-row
shape does not match, so every 2013-2016 equity row is silently dropped
by the `len(vals) < 7` check - not mis-parsed, just never reaches the
data-row branch at all. Combined with the already-known 2016-05 to
2023-01 filing gap (no files exist for that window at all), the honest
picture is: SBI's real per-stock holdings coverage here is 2023-2026
only; a legacy-template parser for 2013-2016 is real, unscoped follow-on
work, not started this session.

PCT_TO_NAV IS NORMALIZED TO A FRACTION (0-1) FOR BOTH AMCS IN THE
COMBINED OUTPUT - a real, confirmed unit difference, not assumed: Axis's
raw sheets already store this as a fraction (e.g. 0.0759 = 7.59%); SBI's
raw sheets store the same quantity as a percentage (e.g. 6.72 meaning
6.72%) - confirmed live by a real scheme's own equity-section sum landing
at 97.99 (i.e. a sane ~98% invested), not ~1.0. SBI's raw values are
divided by 100 before writing; Axis's are left as-is.
"""

from __future__ import annotations

import io
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import xlrd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dtest import load_config
from dtest.data.amc_holdings import SCHEMA, parse_scheme_sheet

CONSOLIDATED_ONLY_PERIOD = pd.Timestamp("2021-09-30")  # the one Axis month per-scheme files don't cover


def _read_rows_by_sheet(path: Path) -> dict[str, list[tuple]]:
    """Real files are either modern XLSX-zip content or legacy BIFF - the
    format must be sniffed from content, never the file's own extension
    (confirmed live: `openpyxl.load_workbook` rejects a real ZIP/XLSX file
    outright if its NAME ends in `.xls`, regardless of actual content - a
    real trap, worked around by always reading bytes first and handing
    `openpyxl` a nameless `io.BytesIO` buffer instead of the path). True
    for both SBI's and Axis's own files - confirmed live on both (SBI: 47
    real ZIP + 21 real OLE2 among its 68 files, extension always matching
    actual content by luck; Axis: extension and content mismatch on many
    files, the case this dispatch was originally built for)."""
    data = path.read_bytes()
    if data[:2] == b"PK":
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        return {name: list(wb[name].iter_rows(values_only=True)) for name in wb.sheetnames}
    wb = xlrd.open_workbook(file_contents=data)
    return {sh.name: [sh.row_values(r) for r in range(sh.nrows)] for sh in wb.sheets()}


_SBI_SCHEME_NAME_LABEL_RE = re.compile(r"^(scheme\s*name|name\s*of\s*the\s*scheme)\s*:?$", re.I)


def _sbi_scheme_index(sheets: dict[str, list[tuple]]) -> dict[str, tuple[str, str]]:
    """SBI's own "Index" sheet: rows of (code, short_code, full_name) after
    a 3-row header/title block - maps short_code (the real sheet name used
    for every OTHER sheet in this workbook) -> (code, full_name). SOME
    OLDER REAL FILES (10 of SBI's 68, all 2013-2015) have a genuinely
    DIFFERENT "Index" sheet shape - a bare list of sheet names with no
    code/full-name columns at all - so this returns an empty mapping for
    those; the caller falls back to `_sbi_scheme_own_identity` per sheet."""
    out = {}
    for row in sheets.get("Index", []):
        vals = [c for c in row if c is not None and str(c).strip() != ""]
        if len(vals) == 3 and str(vals[0]).strip() != "Scheme Code":
            code, short_code, name = (str(v).strip() for v in vals)
            out[short_code] = (code, name)
    return out


def _sbi_scheme_own_identity(rows: list[tuple]) -> tuple[str | None, str | None]:
    """Fallback for the 10 older files `_sbi_scheme_index` can't map: reads
    the scheme's OWN identity rows directly - row 0 = [AMC name, code] or
    [AMC name, code, "Back to Index"], and a labeled row ("SCHEME NAME :"
    or the older "NAME OF THE SCHEME :") pairs with the real full name.
    Both label wordings and both row-0 shapes confirmed real, not assumed
    from one era."""
    code = name = None
    for row in rows[:6]:
        vals = [c for c in row if c is not None and str(c).strip() != ""]
        if len(vals) < 2:
            continue
        label = str(vals[0]).strip()
        if name is None and _SBI_SCHEME_NAME_LABEL_RE.match(label):
            name = str(vals[1]).strip()
        elif code is None and not _SBI_SCHEME_NAME_LABEL_RE.match(label):
            # row 0 candidate: [AMC name, code, ...] - code is the 2nd cell
            candidate = str(vals[1]).strip()
            if len(candidate) <= 10:  # a real scheme code, not a stray long value
                code = candidate
        if code is not None and name is not None:
            break
    return code, name


def _holdings_to_rows(amc_name: str, scheme_name: str, scheme_code: str | None,
                       period_end, holdings, pct_scale: float) -> list[dict]:
    return [{
        "amc_name": amc_name, "scheme_name": scheme_name, "scheme_code": scheme_code,
        "period_end": period_end, "instrument_code": h.instrument_code,
        "instrument_name": h.instrument_name, "isin": h.isin, "industry": h.industry,
        "quantity": h.quantity, "market_value_lakhs": h.market_value_lakhs,
        "pct_to_nav": h.pct_to_nav / pct_scale,
    } for h in holdings]


def build_axis(amc_dir: Path) -> list[dict]:
    manifest = pd.read_csv(amc_dir / "manifest.csv", parse_dates=["period_end"])
    axis = manifest[manifest["amc_name"] == "Axis Mutual Fund"]

    per_scheme = axis[axis["scheme_name"].notna()].copy()
    # REAL DEDUP, NOT OPTIONAL: `amc_portfolios.py`'s own module docstring
    # already documents 13 (scheme, period_end) collisions where two
    # DIFFERENT real filing-metadata entries (different title casing, e.g.
    # "Axis Nifty Smallcap 50 Index Fund" vs "AXIS NIFTY SMALLCAP 50 INDEX
    # FUND") resolve to the SAME local filename - Windows filesystems are
    # case-insensitive, so both manifest rows point at one physical file on
    # disk. Left unfixed here, this loop would parse that one file twice
    # under two scheme_name labels, silently doubling every holding for
    # those pairs (caught live: several schemes' own pct_to_nav summed to
    # ~2.0 instead of the normal ~0.85-1.0 before this fix).
    n_before = len(per_scheme)
    per_scheme["_dedup_key"] = per_scheme["local_filename"].str.lower()
    per_scheme = per_scheme.drop_duplicates(subset="_dedup_key", keep="first")
    print(f"[axis] deduped {n_before - len(per_scheme)} manifest rows sharing a "
          f"case-insensitive-identical local_filename with another row")

    consolidated_only_row = axis[(axis["scheme_name"].isna())
                                  & (axis["period_end"] == CONSOLIDATED_ONLY_PERIOD)]

    rows: list[dict] = []
    n_ok, n_failed, n_date_mismatch = 0, 0, 0

    for _, filing in per_scheme.iterrows():
        path = amc_dir / filing["local_filename"]
        try:
            sheets = _read_rows_by_sheet(path)
            sheet_rows = next(iter(sheets.values()))
        except Exception as e:
            print(f"  FAILED {filing['local_filename']}: {e}")
            n_failed += 1
            continue
        holdings = parse_scheme_sheet(sheet_rows)
        n_ok += 1
        for h in holdings:
            if h.sheet_as_on_date is not None and h.sheet_as_on_date != filing["period_end"]:
                n_date_mismatch += 1
        rows.extend(_holdings_to_rows("Axis Mutual Fund", filing["scheme_name"], None,
                                       filing["period_end"], holdings, pct_scale=1.0))
        if n_ok % 500 == 0:
            print(f"  [axis] {n_ok}/{len(per_scheme)} per-scheme files processed, {len(rows)} equity rows so far")

    for _, filing in consolidated_only_row.iterrows():
        path = amc_dir / filing["local_filename"]
        sheets = _read_rows_by_sheet(path)
        for sheet_name, sheet_rows in sheets.items():
            if sheet_name.lower() == "index":
                continue
            holdings = parse_scheme_sheet(sheet_rows)
            for h in holdings:
                rows.extend(_holdings_to_rows("Axis Mutual Fund", h.scheme_name, None,
                                               filing["period_end"], [h], pct_scale=1.0))
        print(f"  [axis] consolidated {filing['local_filename']}: {len(sheets)-1} scheme sheets processed")

    print(f"[axis] {n_ok}/{len(per_scheme)} per-scheme files parsed ({n_failed} failed), "
          f"{n_date_mismatch} rows with a sheet-vs-filename date mismatch")
    return rows


def build_sbi(amc_dir: Path) -> list[dict]:
    manifest = pd.read_csv(amc_dir / "manifest.csv", parse_dates=["period_end"])
    sbi = manifest[manifest["amc_name"] == "SBI Mutual Fund"]

    rows: list[dict] = []
    n_files, n_schemes, n_failed = 0, 0, 0

    for _, filing in sbi.iterrows():
        n_files += 1
        path = amc_dir / filing["local_filename"]
        try:
            sheets = _read_rows_by_sheet(path)
        except Exception as e:
            print(f"  FAILED {filing['local_filename']}: {e}")
            n_failed += 1
            continue
        scheme_index = _sbi_scheme_index(sheets)
        n_this_file = 0
        for sheet_name, sheet_rows in sheets.items():
            if sheet_name.lower() in ("index", "tamplate", "template"):
                continue
            if sheet_name in scheme_index:
                code, name = scheme_index[sheet_name]
            else:
                # ~10 older real files have a bare Index sheet with no
                # code/full-name columns at all (see `_sbi_scheme_index`'s
                # own docstring) - fall back to that scheme's own identity
                # rows rather than dropping the whole file's real data.
                code, name = _sbi_scheme_own_identity(sheet_rows)
                if name is None:
                    continue
                code = code or sheet_name
            holdings = parse_scheme_sheet(sheet_rows, scheme_code=code, scheme_name=name)
            n_schemes += 1
            n_this_file += 1
            # SBI stores pct_to_nav as a PERCENTAGE (e.g. 6.72 = 6.72%),
            # confirmed real - divide by 100 to match Axis's fraction scale.
            rows.extend(_holdings_to_rows("SBI Mutual Fund", name, code,
                                           filing["period_end"], holdings, pct_scale=100.0))
        print(f"  [sbi] {filing['local_filename']}: {n_this_file} scheme sheets processed")

    print(f"[sbi] {n_files-n_failed}/{n_files} workbooks parsed ({n_failed} failed), "
          f"{n_schemes} scheme-months extracted")
    return rows


def main() -> int:
    cfg = load_config()
    amc_dir = Path(cfg.paths.amc_portfolios_dir)

    all_rows = build_axis(amc_dir) + build_sbi(amc_dir)
    if not all_rows:
        print("No holdings extracted from any AMC - nothing to write.")
        return 1

    df = pd.DataFrame(all_rows)
    df = df.astype({k: v for k, v in SCHEMA.items() if k in df.columns and k != "sheet_as_on_date"})
    df = df.sort_values(["amc_name", "scheme_name", "period_end", "pct_to_nav"],
                         ascending=[True, True, True, False]).reset_index(drop=True)
    out_path = amc_dir / "equity_holdings.csv"
    df.to_csv(out_path, index=False)
    print(f"\nWrote {len(df)} equity-holding rows ({df['amc_name'].nunique()} AMCs, "
          f"{df['scheme_name'].nunique()} schemes, {df['period_end'].nunique()} distinct months) "
          f"to {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
